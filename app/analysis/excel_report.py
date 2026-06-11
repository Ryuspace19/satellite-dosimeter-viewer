from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


SHEET_MAP = {
    "timeseries": "全データ_時系列整理",
    "temperature_fits": "全期間_Vth温度フィット",
    "daily": "日ごと統計",
    "weekly": "週ごと統計",
    "binned": "任意期間統計",
    "time_trends": "時間トレンド判定",
}


def write_excel_report(results: dict[str, pd.DataFrame], output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        explanation = pd.DataFrame(
            [
                ["目的", "衛星線量計の復号済みVthデータを温度補正・期間集計し、時間トレンドを評価する。"],
                ["主指標", "mos1_vth_interp, mos2_vth_interp, mos_mean_vth25"],
                ["温度補正", "Vth25 = Vth_interp - A_GROUND * (temp_degC - 25)"],
                ["線量換算", "18.5 mV/kradを使用。絶対値でrad換算する。"],
                ["注意", "rawデータは変更せず、解析結果のみを本ブックへ出力する。"],
            ],
            columns=["項目", "説明"],
        )
        explanation.to_excel(writer, sheet_name="説明", index=False)
        for key, sheet_name in SHEET_MAP.items():
            results[key].to_excel(writer, sheet_name=sheet_name, index=False)

        workbook = writer.book
        header_fill = PatternFill("solid", fgColor="1F4E78")
        for sheet in workbook.worksheets:
            sheet.freeze_panes = "A2"
            sheet.auto_filter.ref = sheet.dimensions
            for cell in sheet[1]:
                cell.fill = header_fill
                cell.font = Font(color="FFFFFF", bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            for column_cells in sheet.columns:
                values = [str(cell.value) if cell.value is not None else "" for cell in column_cells[:200]]
                width = min(max(max((len(value) for value in values), default=8) + 2, 10), 42)
                sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = width
